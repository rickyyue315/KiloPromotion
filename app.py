import streamlit as st
import tempfile
import os
from datetime import datetime
import io
import requests
import base64
from logger import logger

def check_for_updates():
    try:
        # Read local version
        with open('VERSION.md', 'r', encoding='utf-8') as f:
            local_content = f.read()
        local_version = local_content.split('\n')[0].replace('# Version ', '')

        # Assume GitHub repo URL
        repo_url = "https://api.github.com/repos/example/repo/contents/VERSION.md"
        response = requests.get(repo_url)
        if response.status_code == 200:
            data = response.json()
            remote_content = base64.b64decode(data['content']).decode('utf-8')
            remote_version = remote_content.split('\n')[0].replace('# Version ', '')
            if remote_version != local_version:
                st.info(f"有新版本可用: {remote_version} (當前: {local_version})")
            else:
                st.success("已是最新版本")
        else:
            st.error("無法檢查更新")
    except Exception as e:
        logger.error(f"Error checking updates: {str(e)}")
        st.error("檢查更新失敗")

# Dependency check
try:
    import pandas as pd
    import numpy as np
    import openpyxl
    from data_preprocessing import load_and_preprocess
    from business_logic import calculate_demand
    from visualization import create_visualizations
except ImportError as e:
    logger.error(f"Import error: {str(e)}")
    st.error("缺少必要套件: " + str(e))
    st.stop()

def create_excel(df_raw, df_results, df_summary):
    wb = openpyxl.Workbook()
    # Sheet 1: Raw Data
    ws1 = wb.active
    ws1.title = "Raw Data"
    for r, row in enumerate([df_raw.columns.tolist()] + df_raw.values.tolist(), 1):
        for c, val in enumerate(row, 1):
            ws1.cell(row=r, column=c, value=val)
    # Sheet 2: Calculation Results
    ws2 = wb.create_sheet("Calculation Results")
    for r, row in enumerate([df_results.columns.tolist()] + df_results.values.tolist(), 1):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val)
    # Sheet 3: Summary
    ws3 = wb.create_sheet("Summary")
    for r, row in enumerate([df_summary.columns.tolist()] + df_summary.values.tolist(), 1):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

# Page configuration
st.set_page_config(
    page_title="零售推廣目標檢視及派貨系統",
    page_icon="📊",
    layout="wide"
)

# Main title
st.title("零售推廣目標檢視及派貨系統")

# Create tabs
tab1, tab2, tab3, tab4 = st.tabs(["📤 數據上傳與分析", "📊 計算結果", "📈 視覺化分析", "📥 匯出報告"])

# Sidebar
with st.sidebar:
    st.header("Developer Info")
    st.write("Developer: Ricky")
    st.write("Version: v1.0")

    if st.button("檢查更新", key="update_check"):
        check_for_updates()

    st.header("參數設定")
    lead_time = st.slider("Lead Time (days)", min_value=2.0, max_value=5.0, value=2.0, step=0.5)

    st.header("檔案格式說明")
    with st.expander("檔案 A (庫存與銷售數據)"):
        st.write("必要欄位: Article, Article Description, RP Type, Site, MOQ, SaSa Net Stock, Pending Received, Safety Stock, Last Month Sold Qty, MTD Sold Qty, Supply source, Description p. group")
        st.write("RP Type: 'ND' 或 'RF'")
        st.write("Supply source: 1=行, 2=倉, 4=行")

    with st.expander("檔案 B (推廣目標數據)"):
        st.write("Sheet1: Group No., Article, SKU Target, Target Type ('HK'/'MO'/'ALL'), Promotion Days, Target Cover Days")
        st.write("Sheet2: Site, Shop Target(HK), Shop Target(MO), Shop Target(ALL)")

# Global variables for data
if 'df_raw' not in st.session_state:
    st.session_state.df_raw = pd.DataFrame()
if 'df_results' not in st.session_state:
    st.session_state.df_results = pd.DataFrame()
if 'summary' not in st.session_state:
    st.session_state.summary = pd.DataFrame()

with tab1:
    st.header("數據上傳與分析")
    file_a = st.file_uploader("上傳檔案 A (庫存與銷售數據)", type=['xlsx'])
    file_b = st.file_uploader("上傳檔案 B (推廣目標數據)", type=['xlsx'])

    if file_a and file_b:
        if st.button("開始分析", key="analyze"):
            # Process in memory
            df_raw = load_and_preprocess(file_a.getvalue(), file_b.getvalue())

            if not df_raw.empty:
                # Data preview
                st.subheader("數據預覽")
                st.dataframe(df_raw.head(10))

                # Progress bar for analysis
                progress_bar = st.progress(0)
                status_text = st.empty()

                # Calculate demand
                status_text.text("正在分析中...")
                progress_bar.progress(50)
                df_results, summary = calculate_demand(df_raw.copy(), lead_time)
                progress_bar.progress(100)
                status_text.text("分析完成！")

                # Store in session state
                st.session_state.df_raw = df_raw
                st.session_state.df_results = df_results
                st.session_state.summary = summary

                st.success("✅ 分析完成！")
                st.rerun()  # Refresh to show other tabs
            else:
                st.error("處理數據失敗。請檢查上方錯誤訊息。")

with tab2:
    st.header("計算結果")
    if not st.session_state.df_results.empty:
        st.subheader("詳細計算結果")
        st.dataframe(st.session_state.df_results, width='stretch')
        st.subheader("總結報告 (按組別與SKU)")
        st.dataframe(st.session_state.summary, width='stretch')
    else:
        st.info("請先上傳檔案並進行分析。")

with tab3:
    st.header("視覺化分析")
    if not st.session_state.df_results.empty:
        create_visualizations(st.session_state.df_results, st.session_state.summary)
    else:
        st.info("請先上傳檔案並進行分析。")

with tab4:
    st.header("匯出報告")
    if not st.session_state.df_raw.empty:
        bio = create_excel(st.session_state.df_raw, st.session_state.df_results, st.session_state.summary)
        date_str = datetime.now().strftime("%Y%m%d")
        filename = f"Promotion_Demand_Report_{date_str}.xlsx"
        st.download_button(
            label="下載 Excel 報告",
            data=bio,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel"
        )
    else:
        st.info("請先上傳檔案並進行分析。")

# Sidebar
with st.sidebar:
    st.header("Developer Info")
    st.write("Developer: Ricky")
    st.write("Version: v1.0")

    if st.button("檢查更新"):
        check_for_updates()

    st.header("File Format Notes")
    with st.expander("File A (Inventory and Sales Data)"):
        st.write("Required columns: Article, Article Description, RP Type, Site, MOQ, SaSa Net Stock, Pending Received, Safety Stock, Last Month Sold Qty, MTD Sold Qty, Supply source, Description p. group")
        st.write("RP Type: 'ND' or 'RF'")
        st.write("Supply source: 1=行, 2=倉, 4=行")

    with st.expander("File B (Promotion Target)"):
        st.write("Sheet1: Group No., Article, SKU Target, Target Type ('HK'/'MO'/'ALL'), Promotion Days, Target Cover Days")
        st.write("Sheet2: Site, Shop Target(HK), Shop Target(MO), Shop Target(ALL)")

    st.header("Quick Navigation")
    # Placeholder for navigation items
    st.write("Upload")
    st.write("Analysis")
    st.write("Charts")